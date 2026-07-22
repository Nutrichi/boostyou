#!/usr/bin/env python3
"""BiS lists <-> Excel workbook.

Export (build/refresh the template, pre-filled with all existing lists):
    python3 tools/bis_excel.py export            -> bis-lists.xlsx
Import (convert a filled workbook into content/bis-data/*.json):
    python3 tools/bis_excel.py import bis-lists.xlsx

Workbook layout (see README sheet inside the file):
  Lists    : expansion | phase | class | spec | role | status | notes
  Items    : expansion | phase | class | spec | slot | rank | item | source
  Enchants : expansion | phase | class | spec | slot | enchant | source
  Gems     : expansion | phase | class | spec | color | gem

Friendly spellings are accepted on import ("Death Knight", "WotLK",
"Phase 4", "pre-raid", ...) and normalized to the ids that
assets/bis.js and the bis-data filenames use.
"""

import datetime
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "content" / "bis-data"
DEFAULT_XLSX = ROOT / "bis-lists.xlsx"

EXPANSIONS = {
    "classic": ["preraid", "p1", "p2", "p3", "p4", "p5", "p6"],
    "tbc": ["preraid", "p1", "p2", "p3", "p4", "p5"],
    "wotlk": ["preraid", "p1", "p2", "p3", "p4"],
    "cata": ["preraid", "p1", "p2", "p3"],
    "mop": ["preraid", "p1", "p2", "p3"],
}

CLASSES = {
    "death-knight": ["blood", "frost", "unholy"],
    "druid": ["balance", "feral", "guardian", "restoration"],
    "hunter": ["beast-mastery", "marksmanship", "survival"],
    "mage": ["arcane", "fire", "frost"],
    "monk": ["brewmaster", "mistweaver", "windwalker"],
    "paladin": ["holy", "protection", "retribution"],
    "priest": ["discipline", "holy", "shadow"],
    "rogue": ["assassination", "combat", "subtlety"],
    "shaman": ["elemental", "enhancement", "restoration"],
    "warlock": ["affliction", "demonology", "destruction"],
    "warrior": ["arms", "fury", "protection"],
}

EXP_ALIASES = {
    "vanilla": "classic", "era": "classic",
    "bc": "tbc", "burning crusade": "tbc", "the burning crusade": "tbc",
    "wrath": "wotlk", "wrath of the lich king": "wotlk",
    "cataclysm": "cata",
    "mists": "mop", "mists of pandaria": "mop",
}

CLASS_ALIASES = {"dk": "death-knight", "deathknight": "death-knight"}

SLOT_ORDER = ["Head", "Neck", "Shoulders", "Back", "Chest", "Wrists", "Hands",
              "Waist", "Legs", "Feet", "Ring 1", "Ring 2", "Trinket 1",
              "Trinket 2", "Weapon", "Main Hand", "Off-Hand", "Two-Hand",
              "Ranged", "Wand", "Idol", "Relic", "Totem", "Libram", "Sigil"]

HEADERS = {
    "Lists": ["expansion", "phase", "class", "spec", "role", "status", "notes"],
    "Items": ["expansion", "phase", "class", "spec", "slot", "rank", "item", "source"],
    "Enchants": ["expansion", "phase", "class", "spec", "slot", "enchant", "source"],
    "Gems": ["expansion", "phase", "class", "spec", "color", "gem"],
}

README = [
    "BOOSTYOU.AI — BiS LIST WORKBOOK",
    "",
    "One BiS list = one (expansion, phase, class, spec) combination.",
    "The 'Items' sheet is the core: one row per item, up to 3 ranked items per slot",
    "(rank 1 = BiS, 2 = second best, 3 = third best).",
    "'source' is free text, e.g.: Icecrown Citadel 25 Heroic — The Lich King",
    "",
    "Sheets:",
    "  Lists    — one row per list: role (Healer/DPS/Tank), status (draft or empty), notes.",
    "  Items    — expansion | phase | class | spec | slot | rank | item | source",
    "  Enchants — one row per slot enchant for that list.",
    "  Gems     — one row per socket color (Meta / Red / Yellow / Blue) for that list.",
    "",
    "Accepted values (friendly spellings like 'WotLK', 'Death Knight', 'Phase 4',",
    "'Pre-Raid' are fine — they get normalized on import):",
    "  expansion: Classic, TBC, WotLK, Cata, MoP",
    "  phase:     Pre-Raid, P1 ... P6 (Classic to P6, TBC to P5, WotLK to P4, Cata/MoP to P3)",
    "  class/spec: the usual — Beast Mastery, Restoration, Mistweaver, ...",
    "  slot:      " + ", ".join(SLOT_ORDER),
    "",
    "Import with:  python3 tools/bis_excel.py import bis-lists.xlsx",
    "Rows with problems are reported with their sheet + row number and skipped.",
]


def norm(value):
    return str(value).strip() if value is not None else ""


def norm_id(value, aliases=None, extra=None):
    v = norm(value).lower().replace("_", " ").replace("-", " ").strip()
    if aliases and v in aliases:
        return aliases[v]
    if extra:
        v = extra(v)
    return v.replace(" ", "-")


def norm_phase(value):
    v = norm(value).lower().replace("-", "").replace(" ", "")
    if v in ("preraid", "praid", "pre"):
        return "preraid"
    m = re.fullmatch(r"(?:p|phase)?(\d)", v)
    return "p" + m.group(1) if m else v


def list_key(exp, phase, cls, spec):
    return f"{exp}-{phase}-{cls}-{spec}"


# ---------------------------------------------------------------- export


def export(path):
    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="3B82F6")

    ws = wb.active
    ws.title = "README"
    for line in README:
        ws.append([line])
    ws.column_dimensions["A"].width = 100

    for name, headers in HEADERS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
        widths = {"notes": 60, "item": 42, "source": 60, "enchant": 46, "gem": 60, "slot": 12}
        for i, h in enumerate(headers):
            ws.column_dimensions[chr(ord("A") + i)].width = widths.get(h, 14)
        ws.freeze_panes = "A2"

    for f in sorted(DATA_DIR.glob("*.json")):
        data = json.loads(f.read_text(encoding="utf-8"))
        ident = [data["expansion"], data["phase"], data["class"], data["spec"]]
        wb["Lists"].append(ident + [data.get("role", ""), data.get("status", ""), data.get("notes", "")])
        for s in data.get("slots", []):
            for rank, item in enumerate(s.get("items", []), start=1):
                wb["Items"].append(ident + [s["slot"], rank, item.get("name", ""), item.get("source", "")])
        for e in data.get("enchants", []):
            wb["Enchants"].append(ident + [e.get("slot", ""), e.get("name", ""), e.get("source", "")])
        for g in data.get("gems", []):
            wb["Gems"].append(ident + [g.get("color", ""), g.get("name", "")])

    wb.save(path)
    print(f"Wrote {path}")


# ---------------------------------------------------------------- import


def read_rows(wb, sheet):
    if sheet not in wb.sheetnames:
        return
    for i, row in enumerate(wb[sheet].iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or norm(v) == "" for v in row):
            continue
        yield i, list(row) + [None] * (len(HEADERS[sheet]) - len(row))


def parse_ident(row, sheet, rownum, errors):
    exp = norm_id(row[0], EXP_ALIASES)
    phase = norm_phase(row[1])
    cls = norm_id(row[2], CLASS_ALIASES)
    spec = norm_id(row[3])
    if exp not in EXPANSIONS:
        errors.append(f"{sheet} row {rownum}: unknown expansion {row[0]!r}")
        return None
    if phase not in EXPANSIONS[exp]:
        errors.append(f"{sheet} row {rownum}: unknown phase {row[1]!r} for {exp}")
        return None
    if cls not in CLASSES:
        errors.append(f"{sheet} row {rownum}: unknown class {row[2]!r}")
        return None
    if spec not in CLASSES[cls]:
        errors.append(f"{sheet} row {rownum}: unknown spec {row[3]!r} for {cls}")
        return None
    return exp, phase, cls, spec


def import_wb(path):
    wb = load_workbook(path, data_only=True)
    errors = []
    lists = {}

    def entry(ident):
        key = list_key(*ident)
        if key not in lists:
            exp, phase, cls, spec = ident
            lists[key] = {
                "expansion": exp, "phase": phase, "class": cls, "spec": spec,
                "role": "", "status": "", "notes": "",
                "updated": datetime.date.today().isoformat(),
                "_slots": {}, "enchants": [], "gems": [],
            }
        return lists[key]

    for rownum, row in read_rows(wb, "Lists"):
        ident = parse_ident(row, "Lists", rownum, errors)
        if not ident:
            continue
        e = entry(ident)
        e["role"], e["status"], e["notes"] = norm(row[4]), norm(row[5]).lower(), norm(row[6])

    for rownum, row in read_rows(wb, "Items"):
        ident = parse_ident(row, "Items", rownum, errors)
        if not ident:
            continue
        slot, rank, item, source = norm(row[4]), norm(row[5]), norm(row[6]), norm(row[7])
        if not slot or not item:
            errors.append(f"Items row {rownum}: slot and item are required")
            continue
        try:
            rank = int(float(rank)) if rank else 1
        except ValueError:
            errors.append(f"Items row {rownum}: rank must be 1, 2 or 3 (got {rank!r})")
            continue
        if rank not in (1, 2, 3):
            errors.append(f"Items row {rownum}: rank must be 1, 2 or 3 (got {rank})")
            continue
        entry(ident)["_slots"].setdefault(slot, []).append((rank, {"name": item, "source": source}))

    for rownum, row in read_rows(wb, "Enchants"):
        ident = parse_ident(row, "Enchants", rownum, errors)
        if not ident:
            continue
        if not norm(row[5]):
            errors.append(f"Enchants row {rownum}: enchant name is required")
            continue
        enchant = {"slot": norm(row[4]), "name": norm(row[5])}
        if norm(row[6]):
            enchant["source"] = norm(row[6])
        entry(ident)["enchants"].append(enchant)

    for rownum, row in read_rows(wb, "Gems"):
        ident = parse_ident(row, "Gems", rownum, errors)
        if not ident:
            continue
        if not norm(row[4]) or not norm(row[5]):
            errors.append(f"Gems row {rownum}: color and gem are required")
            continue
        entry(ident)["gems"].append({"color": norm(row[4]).title(), "name": norm(row[5])})

    for err in errors:
        print("SKIPPED:", err)

    written = 0
    for key, data in sorted(lists.items()):
        slots = data.pop("_slots")
        if not slots:
            print(f"SKIPPED: {key} — no item rows")
            continue
        def slot_pos(name):
            return SLOT_ORDER.index(name) if name in SLOT_ORDER else len(SLOT_ORDER)
        data["slots"] = [
            {"slot": s, "items": [it for _, it in sorted(slots[s], key=lambda r: r[0])]}
            for s in sorted(slots, key=slot_pos)
        ]
        for field in ("role", "status", "notes"):
            if not data[field]:
                del data[field]
        data["gems"] = data.pop("gems")  # keep key order: slots, enchants, gems
        out = DATA_DIR / f"{key}.json"
        out.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(f"Wrote {out.relative_to(ROOT)}  ({len(data['slots'])} slots, "
              f"{len(data.get('enchants', []))} enchants, {len(data.get('gems', []))} gems)")
        written += 1

    print(f"\n{written} list(s) written, {len(errors)} problem row(s) skipped.")
    return 1 if errors and not written else 0


def main():
    args = sys.argv[1:]
    if args and args[0] == "export":
        export(Path(args[1]) if len(args) > 1 else DEFAULT_XLSX)
    elif args and args[0] == "import":
        if len(args) < 2:
            sys.exit("Usage: bis_excel.py import <workbook.xlsx>")
        sys.exit(import_wb(Path(args[1])))
    else:
        sys.exit(__doc__)


if __name__ == "__main__":
    main()
